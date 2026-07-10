"""
Fills template/blank_consolidated_quote_template.xlsx ("Raw Quote" sheet) from
the JSON produced by consolidated_gemini_client.extract_consolidated_bills()
— a list of per-site bills instead of excel_filler.py's single bill.

TEMPLATE STRUCTURE:

    Row 18-20   Fixed table header (site columns + "Current Energy Offer" /
                "New Proposed Offer" / "% Savings" / "Key Feature" banners).
                Never moves, never repeats.

    Row 21+     One BLOCK per site, made of exactly three PARTS, in order:
                  - N "charge line" rows (one per real charge on the bill;
                    the template ships an example with 2, but this is
                    driven entirely by len(bill["charges"]))
                  - exactly ONE blank "spacer" row, always present,
                    directly above the Total row
                  - 1 Total row
                Columns A-F (site address/distributor/NMI/tariff/OC number/
                retailer), G (annual usage), S (% savings), T (key feature)
                are each a single cell MERGED across the whole block — one
                value per site, not per charge line.

    Below the   "Estimated Annual Costs" / "Estimated Annual Savings" summary
    blocks      rows, which roll up EVERY site's total (each weighted by that
                site's own billing period), then the "KEY FEATURES" box and
                the static footer/T&Cs text — identical wording to the
                single-site template.

The shipped reference file has exactly 3 example site-blocks, each with a
visibly different hand-finished style (the first block's top border ties
into the header, the last block's bottom border closes the table, the
middle block is plain). This module treats those three as FIRST / MIDDLE /
LAST block style archetypes and re-stamps whichever one applies to each
site — so a 5-site quote looks like block1, block2, block2, block2, block3,
and a 1-site quote looks like block1 with block3's closing border patched
onto its bottom row.

Each archetype is captured as exactly 4 "roles", by role rather than by
fixed relative row number, since the number of charge rows varies per site:
    role 0 ("first charge")  <- template's 1st charge row
    role 1 ("later charge")  <- template's 2nd charge row (repeated for
                                 every charge after the first, however many
                                 there are)
    role 2 ("spacer")        <- template's 3rd row (blank in the shipped
                                 example) — always exactly one row, always
                                 blank, always directly above Total
    role 3 ("total")         <- template's 4th / Total row
Each role also carries its own row height, which is re-stamped along with
cell styles — this is required because openpyxl's insert_rows() does NOT
carry row heights (or styles) onto newly created rows, it only shifts
existing rows down; without re-stamping, every inserted row silently falls
back to the sheet's default row height and the block looks compressed.

Row-count changes needed for a given list of bills happen in two layered
steps, each reusing the same insert/delete-with-merge-fix approach
excel_filler.py uses for the single-site template:

  1. Normalize the NUMBER OF BLOCKS to match len(bills) (insert or delete
     whole blocks immediately after the last block / before the summary
     section).
  2. Walk the blocks top-to-bottom; any site whose (charges + spacer) needs
     more than the reserved 3 rows gets extra rows inserted inside its own
     block (shifting every block below it, and the summary section, down
     by that amount). Because this walk is top-to-bottom with a running
     offset, insertions in an earlier block never invalidate a later
     block's already-computed position.
"""
import copy
import os

import openpyxl
from openpyxl.worksheet.cell_range import CellRange
from collections import Counter

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "template", "blank_consolidated_quote_template.xlsx")
SHEET_NAME = "Raw Quote"

FIRST_BLOCK_START = 21          # first site block's first charge row
DEFAULT_CHARGE_ROWS = 2         # charge rows shown in the shipped example
RESERVED_ROWS_BEFORE_TOTAL = DEFAULT_CHARGE_ROWS + 1   # + spacer row = 3
BLOCK_HEIGHT = RESERVED_ROWS_BEFORE_TOTAL + 1           # + Total row = 4
TEMPLATE_BLOCK_COUNT = 3        # how many example blocks ship in the template

STYLE_COLS = [chr(c) for c in range(ord("A"), ord("U") + 1)]  # sheet uses up to column U

# Columns merged once across an entire site block (one value per site).
BLOCK_MERGED_COLS = ["A", "B", "C", "D", "E", "F", "G", "N", "S", "T"]

# Columns that hold per-charge-row data and must be blanked on unused rows.
CHARGE_DATA_COLS = ["H", "I", "J", "K", "L", "M", "O", "P", "Q", "R"]


# ---------------------------------------------------------------------------
# Generic row insert/delete helpers (same approach as excel_filler.py's
# _insert_rows_preserving_layout: openpyxl's insert_rows/delete_rows moves
# cell values but leaves row_dimensions heights and merged_cells ranges
# behind, so both have to be fixed up by hand).
# ---------------------------------------------------------------------------

def _insert_rows_preserving_layout(ws, insert_at: int, n: int):
    if n <= 0:
        return
    old_heights = {
        r: ws.row_dimensions[r].height
        for r in range(1, ws.max_row + 1)
        if ws.row_dimensions[r].height is not None
    }
    old_merges = [str(mc) for mc in ws.merged_cells.ranges]
    for mc in old_merges:
        ws.unmerge_cells(mc)

    ws.insert_rows(insert_at, n)

    for old_row, height in old_heights.items():
        new_row = old_row + n if old_row >= insert_at else old_row
        ws.row_dimensions[new_row].height = height

    for mc in old_merges:
        cell_range = CellRange(mc)
        min_r, max_r = cell_range.min_row, cell_range.max_row
        min_c, max_c = cell_range.min_col, cell_range.max_col
        if min_r >= insert_at:
            min_r, max_r = min_r + n, max_r + n
        elif max_r >= insert_at:
            max_r += n
        ws.merge_cells(start_row=min_r, start_column=min_c, end_row=max_r, end_column=max_c)


def _delete_rows_preserving_layout(ws, delete_at: int, n: int):
    """
    Deletes rows [delete_at, delete_at+n-1] and fixes merges/heights.
    Only used to trim whole, self-contained trailing blocks, so merges are
    assumed to either sit fully inside the deleted range (dropped) or fully
    outside it (kept, shifted if below) — no partial-overlap handling.
    """
    if n <= 0:
        return
    delete_end = delete_at + n - 1
    old_heights = {
        r: ws.row_dimensions[r].height
        for r in range(1, ws.max_row + 1)
        if ws.row_dimensions[r].height is not None
    }
    old_merges = [str(mc) for mc in ws.merged_cells.ranges]
    for mc in old_merges:
        ws.unmerge_cells(mc)

    ws.delete_rows(delete_at, n)

    for old_row, height in old_heights.items():
        if delete_at <= old_row <= delete_end:
            continue
        new_row = old_row - n if old_row > delete_end else old_row
        ws.row_dimensions[new_row].height = height

    for mc in old_merges:
        cell_range = CellRange(mc)
        min_r, max_r = cell_range.min_row, cell_range.max_row
        min_c, max_c = cell_range.min_col, cell_range.max_col
        if max_r < delete_at:
            pass  # entirely above, unaffected
        elif min_r > delete_end:
            min_r, max_r = min_r - n, max_r - n
        else:
            continue  # entirely (or mostly) inside the deleted block — drop it
        ws.merge_cells(start_row=min_r, start_column=min_c, end_row=max_r, end_column=max_c)


# ---------------------------------------------------------------------------
# Style archetype capture — snapshot the 3 example blocks' per-cell styling
# (AND row heights) BEFORE any structural changes, so both can be re-stamped
# afterwards regardless of how many blocks/rows the final sheet ends up with.
# ---------------------------------------------------------------------------

def _cell_style_snapshot(cell):
    return {
        "font": copy.copy(cell.font),
        "fill": copy.copy(cell.fill),
        "border": copy.copy(cell.border),
        "alignment": copy.copy(cell.alignment),
        "number_format": cell.number_format,
    }


def _capture_block_archetype(ws, block_start_row: int) -> dict:
    """
    Returns {"styles": [{col: style_snapshot} for each of the 4 roles],
              "heights": [height for each of the 4 roles]}
    Role order: 0 = first charge row, 1 = later charge row (template's 2nd
    charge row), 2 = spacer row (template's 3rd, blank row), 3 = Total row.
    """
    styles, heights = [], []
    for offset in range(BLOCK_HEIGHT):
        row = block_start_row + offset
        styles.append({col: _cell_style_snapshot(ws[f"{col}{row}"]) for col in STYLE_COLS})
        heights.append(ws.row_dimensions[row].height)
    return {"styles": styles, "heights": heights}


def _apply_style_snapshot(cell, snapshot):
    cell.font = copy.copy(snapshot["font"])
    cell.fill = copy.copy(snapshot["fill"])
    cell.border = copy.copy(snapshot["border"])
    cell.alignment = copy.copy(snapshot["alignment"])
    cell.number_format = snapshot["number_format"]


def _stamp_row(ws, row: int, style_role: dict, height):
    for col in STYLE_COLS:
        _apply_style_snapshot(ws[f"{col}{row}"], style_role[col])
    ws.row_dimensions[row].height = height


def _apply_block_style(ws, archetype: dict, block_start_row: int, num_charges: int):
    """
    Stamps an archetype onto a block with `num_charges` real charge rows,
    followed by the mandatory blank spacer row, followed by Total — however
    many charge rows that turns out to be (1, 2, 6, whatever).
    """
    styles, heights = archetype["styles"], archetype["heights"]

    # First charge row.
    _stamp_row(ws, block_start_row, styles[0], heights[0])

    # Every charge row after the first reuses the "later charge" role.
    for offset in range(1, num_charges):
        _stamp_row(ws, block_start_row + offset, styles[1], heights[1])

    # Spacer row — always exactly one, always blank, always role 2.
    spacer_row = block_start_row + num_charges
    _stamp_row(ws, spacer_row, styles[2], heights[2])
    for col in CHARGE_DATA_COLS:
        ws[f"{col}{spacer_row}"] = None

    # Total row — always role 3.
    total_row = spacer_row + 1
    _stamp_row(ws, total_row, styles[3], heights[3])


def _reapply_block_merges(ws, block_start_row: int, total_row: int):
    """(Re)creates the per-block merges (A-G, S, T) spanning the block's
    actual current extent, which may differ from the original 4 rows if
    extra charge rows were inserted."""
    col_indices = {col: openpyxl.utils.column_index_from_string(col) for col in BLOCK_MERGED_COLS}
    # Unmerge anything already covering these columns within the block's
    # row range, so re-merging doesn't collide with a stale smaller range.
    for mc in list(ws.merged_cells.ranges):
        if mc.min_row >= block_start_row and mc.max_row <= total_row and mc.min_col in col_indices.values():
            ws.unmerge_cells(str(mc))
    for col in BLOCK_MERGED_COLS:
        ws.merge_cells(f"{col}{block_start_row}:{col}{total_row}")


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------
   

def fill_consolidated_quote(bills_data: list, output_path: str, client_name: str = None) -> str:
    """
    bills_data: list of bill dicts (same shape as excel_filler.fill_quote's
                bill_data, one per site) — typically
                extract_consolidated_bills(pdf_path)["data"]["bills"].
    client_name: overall name to show in "Quote Prepared For" — pass this in
                 explicitly since, for Strata portfolios, no single site's
                 bill actually carries the overall client's name (see
                 oc_number in schema.py).
    """
    if not bills_data:
        raise ValueError("bills_data is empty — need at least one site to build a consolidated quote.")
    
    # Find the most frequent customer_name (ignoring nulls)
    names = [b.get("customer_name") for b in bills_data if b.get("customer_name")]
    most_frequent_name = Counter(names).most_common(1)[0][0] if names else client_name

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb[SHEET_NAME]

    # Set the portfolio-level client name in A6
    ws["A6"] = (most_frequent_name or client_name or "VALUED CUSTOMER").upper()

    if client_name:
        ws["A6"] = client_name

    # --- capture the 3 example blocks' styles + heights before touching row layout ---
    first_archetype = _capture_block_archetype(ws, FIRST_BLOCK_START)
    middle_archetype = _capture_block_archetype(ws, FIRST_BLOCK_START + BLOCK_HEIGHT)
    last_archetype = _capture_block_archetype(ws, FIRST_BLOCK_START + 2 * BLOCK_HEIGHT)

    n = len(bills_data)

    # --- Step 1: normalize block COUNT to match n ---
    if n > TEMPLATE_BLOCK_COUNT:
        extra_blocks = n - TEMPLATE_BLOCK_COUNT
        insert_at = FIRST_BLOCK_START + TEMPLATE_BLOCK_COUNT * BLOCK_HEIGHT
        _insert_rows_preserving_layout(ws, insert_at, extra_blocks * BLOCK_HEIGHT)
        # New blank blocks get the middle style for now; final first/middle/last
        # assignment happens in the per-block pass below regardless.
        for b in range(extra_blocks):
            block_row = insert_at + b * BLOCK_HEIGHT
            _apply_block_style(ws, middle_archetype, block_row, DEFAULT_CHARGE_ROWS)
            _reapply_block_merges(ws, block_row, block_row + DEFAULT_CHARGE_ROWS + 1)
    elif n < TEMPLATE_BLOCK_COUNT:
        removed_blocks = TEMPLATE_BLOCK_COUNT - n
        delete_at = FIRST_BLOCK_START + n * BLOCK_HEIGHT
        _delete_rows_preserving_layout(ws, delete_at, removed_blocks * BLOCK_HEIGHT)

    # --- Step 2: walk blocks top-to-bottom, expanding any block whose site
    #     needs more than the reserved 3 rows (charges + spacer), tracking a
    #     running row offset ---
    row_offset = 0
    block_positions = []  # (block_start_row, total_row, billing_days) per site, in order

    for i, bill in enumerate(bills_data):
        block_start = FIRST_BLOCK_START + i * BLOCK_HEIGHT + row_offset
        charges = bill.get("charges") or []
        num_charges = max(len(charges), 1)
        rows_before_total = num_charges + 1  # charges + mandatory spacer row
        extra = rows_before_total - RESERVED_ROWS_BEFORE_TOTAL

        if extra > 0:
            # Insert the extra charge rows right after the reserved default
            # charge rows, before the (still-reserved) spacer/Total rows.
            insert_at = block_start + DEFAULT_CHARGE_ROWS
            _insert_rows_preserving_layout(ws, insert_at, extra)
            row_offset += extra

        total_row = block_start + rows_before_total
        block_positions.append((block_start, total_row, bill.get("billing_period_days") or 30))

        # --- style + merges for this block's actual extent ---
        if i == 0:
            archetype = first_archetype
        elif i == n - 1:
            archetype = last_archetype
        else:
            archetype = middle_archetype
        _apply_block_style(ws, archetype, block_start, num_charges)
        _reapply_block_merges(ws, block_start, total_row)

        # --- site-level fields (merged across the whole block) ---
        ws[f"A{block_start}"] = bill.get("site_address")
        ws[f"B{block_start}"] = bill.get("distribution_region")
        nmi = bill.get("nmi_or_mirn")
        # ws[f"C{block_start}"].number_format = "@"  # keep as text: NMIs can have leading zeros
        ws[f"C{block_start}"] = str(nmi) if nmi else None
        # ws[f"D{block_start}"] = bill.get("tariff_classification")
        ws[f"E{block_start}"] = bill.get("oc_number")
        ws[f"F{block_start}"] = bill.get("current_energy_retailer")
        oc = bill.get("oc_number")        
        ws[f"E{block_start}"] = oc if oc else ("")
        
        

        # --- charge line rows ---
        for j, charge in enumerate(charges):
            r = block_start + j
            ws[f"H{r}"] = charge.get("quantity")
            desc = charge.get("description") or ""
            if charge.get("is_credit"):
                desc = f"{desc} (credit)" if desc else "Credit"
            ws[f"I{r}"] = desc
            rate = charge.get("rate_before_discount")
            ws[f"J{r}"] = -abs(rate) if (rate is not None and charge.get("is_credit")) else rate
            ws[f"K{r}"] = charge.get("conditional_discount_pct")
            ws[f"L{r}"] = f"=J{r}*(1-K{r})"
            ws[f"M{r}"] = f"=L{r}*H{r}"
            ws[f"Q{r}"] = f"=O{r}*(1-P{r})"
            ws[f"R{r}"] = f"=Q{r}*H{r}"

        # Blank out any leftover reserved charge row (e.g. only 1 real charge
        # but 2 default charge slots) between the last real charge and the
        # spacer row.
        last_charge_row = block_start + num_charges - 1
        for r in range(block_start + len(charges), last_charge_row + 1):
            for col in CHARGE_DATA_COLS:
                ws[f"{col}{r}"] = None
        # (The spacer row itself is already blanked inside _apply_block_style.)

        # --- per-block roll-up formulas ---
        ws[f"I{total_row}"] = "Total (GST Incl.)"
        ws[f"M{total_row}"] = f"=SUM(M{block_start}:M{total_row - 1})"
        ws[f"R{total_row}"] = f"=SUM(R{block_start}:R{total_row - 1})"
        ws[f"S{block_start}"] = f"=1-(R{total_row}/M{total_row})"

        billing_days = bill.get("billing_period_days") or 30
        
        # Collect only the quantities where Gemini marked the charge as 'is_usage'
        usage_charges = [
            c.get("quantity") for c in charges 
            if c.get("is_usage") is True
        ]
        
        # Now we don't have to guess if it's usage or not!
        # Just sum the quantities identified by the AI.
        ws[f"G{block_start}"] = (
            f"=(SUM({' + '.join([f'H{block_start + i}' for i, c in enumerate(charges) if c.get('is_usage')])}) "
            f"/ {billing_days} * 365) / 1000"
        )

    # --- Step 3: rebuild the summary section formulas below all blocks ---
    # These live at fixed OFFSETS below the template's original last block
    # (row 32 -> row 37 is a +5 offset, row 32 -> row 43 is a +11 offset),
    # which stays true regardless of how many blocks/rows precede them since
    # every insertion above this point has already been accounted for.
    original_last_block_end = FIRST_BLOCK_START + TEMPLATE_BLOCK_COUNT * BLOCK_HEIGHT - 1  # 32
    actual_last_block_end = block_positions[-1][1]
    total_offset = actual_last_block_end - original_last_block_end

    annual_costs_row = 37 + total_offset
    annual_savings_row = 43 + total_offset

    current_terms = "+".join(
        f"(M{total_row}/{billing_days})" for (_, total_row, billing_days) in block_positions
    )
    new_terms = "+".join(
        f"(R{total_row}/{billing_days})" for (_, total_row, billing_days) in block_positions
    )
    ws[f"L{annual_costs_row}"] = f"=({current_terms})*365"
    ws[f"Q{annual_costs_row}"] = f"=({new_terms})*365"
    ws[f"L{annual_savings_row}"] = f"=1-(Q{annual_costs_row}/L{annual_costs_row})"
    ws[f"Q{annual_savings_row}"] = f"=L{annual_costs_row}-Q{annual_costs_row}"

    wb.save(output_path)
    return output_path
