"""
retailer_rates.py
------------------
Given a live-looked-up Network Tariff Code (NTC) plus the state/distribution
and charge lines already extracted from a bill, this module finds the
retailer + rate-card combination in Master_Retailers_Comparison.xlsx
("Comparison" tab) that produces the LOWEST total cost for THIS customer's
actual usage, and returns a `proposed_charges` list shaped exactly like
`bill_data["charges"]` (same order, same length) so it can be passed
straight into:

    single_excel_filler.fill_quote(bill_data, out_path,
                                    tariff_override=ntc,
                                    proposed_charges=proposed_charges)

...to populate the "New Proposed Offer" (J/K) columns without touching any
of the row-insertion / merge / formula logic that already works for the
"Current Energy Offer" (E/F) columns.

HOW THE COMPARISON SHEET IS LAID OUT
("Comparison" tab of Master_Retailers_Comparison.xlsx):

    Col A-F  : Type, Tariff Code, CL Component(s), State, Distribution,
               # Retailers Offering
    Col G-T  : ORIGIN      block (14 cols)
    Col U-AH : MOMENTUM    block
    Col AI-AV: 1ST ENERGY  block
    Col AW-BJ: NBE         block
    Col BK-BX: ALINTA      block
    Col BY-CL: EA          block

Each 14-column block, in order:
    Daily supply charge, Peak, Peak 2, Shoulder, Off Peak, CL1, CL2, CL3,
    Capacity Charges, Demand 1, Demand 2, Solar, Discount, Sign-up Credit

A blank cell means that retailer doesn't publish a rate for that field on
that tariff row - this is NOT the same as a rate of $0, and is treated as
"can't fully price this retailer" below (see `_missing`).

ASSUMPTIONS - FLAGGED HERE BECAUSE THEY MATERIALLY CHANGE THE NUMBERS AND
ARE EASY TO GET WRONG SILENTLY. Worth spot-checking against 2-3 known real
offers before trusting this for a live quote:

  1. UNITS: the Comparison sheet's rate fields (everything except Discount
     and Sign-up Credit) are stored in cents and converted to dollars at
     load time (see CENTS_TO_DOLLARS_FIELDS) to match the bill's own
     extracted rates and the template's Current Offer column. Discount is
     read as a plain percentage (20 means 20%) and Sign-up Credit is
     assumed to already be a dollar amount - if either of those turns out
     to be stored differently for some rows, this will misprice just that
     row.
  2. DISCOUNT SCOPE: the same discount fraction is applied to every matched
     charge line, including the Daily Supply Charge, mirroring how the
     template already treats discount per-row (E/F, J/K). If a real offer's
     discount is usage-only, this slightly overstates that retailer's
     saving.
  3. SIGN-UP CREDIT: this is a one-off bill credit, not a per-usage rate, so
     it doesn't fit the line-item J/K structure. It is NOT applied to the
     ranking or written anywhere on the quote - it's returned in
     `meta["signup_credit"]` for Raj to surface manually. Two retailers
     that are ~equal on running cost won't be re-ranked just because one
     has a bigger sign-up credit.
  4. CONTROLLED LOAD: the "CL Component(s)" column (e.g. "060, 070") hints
     that a Base Tariff row is meant to be paired with a specific CL add-on
     tariff row for full accuracy. This module reads CL1/CL2/CL3 straight
     off the SAME row as the base tariff (whatever each retailer publishes
     there) and does NOT separately look up a "Controlled Load Add-on" type
     row. Flag any CL-heavy bill for a manual check.
  5. MATCHING: Tariff Code + State + Distribution (all three, case/space
     insensitive). If nothing matches on all three, falls back to Tariff
     Code + State, then Tariff Code alone - `meta["match_level"]` tells you
     which one actually fired so a loose match is never silent.
  6. UNMATCHED CHARGE LINES: a bill charge whose description can't be
     mapped to one of the 12 rate fields (see `classify_charge`) is left
     out of both the ranking and the proposed_charges output for that row
     (rate=None) - it's flagged in `meta["unmatched_descriptions"]` rather
     than guessed at.
"""
from __future__ import annotations

import os
import re
from dataclasses import dataclass, field

import openpyxl

COMPARISON_SHEET = "Comparison"

# 14 fields per retailer block, in on-sheet column order.
FIELD_ORDER = [
    "daily_supply", "peak", "peak2", "shoulder", "offpeak",
    "cl1", "cl2", "cl3", "capacity", "demand1", "demand2",
    "solar", "discount", "signup_credit",
]

# 1-indexed starting column of each retailer's 14-col block (col G = 7).
RETAILER_START_COL = {
    "ORIGIN": 7,
    "MOMENTUM": 21,
    "1ST ENERGY": 35,
    "NBE": 49,
    "ALINTA": 63,
    "EA": 77,
}

# Fields that represent a credit paid TO the customer rather than a charge.
CREDIT_FIELDS = {"solar"}

# Fields the ranking/costing logic actually uses per bill line (discount and
# signup_credit are handled separately, not matched against a charge line).
CHARGE_FIELDS = [f for f in FIELD_ORDER if f not in ("discount", "signup_credit")]

# The Comparison sheet stores every rate field in CENTS (e.g. 50.00 = 50c),
# same as most AU retail rate cards, but the bill's own extracted rates -
# and the template's Current Offer column - are in DOLLARS (e.g. 0.50).
# Every field in CHARGE_FIELDS is divided by 100 at load time below so the
# New Proposed Offer lines up with the Current Offer's units. Discount (a
# percentage) and Sign-up Credit (already a dollar amount) are left alone.
CENTS_TO_DOLLARS_FIELDS = set(CHARGE_FIELDS)

# Brand colors for the divider column (I30, merged I30:I{total_row-1}) so
# the quote visually flags which retailer the New Proposed Offer is from.
# ARGB hex, as required by openpyxl's PatternFill.
RETAILER_COLORS = {
    "MOMENTUM": "FFADD8E6",     # Light Blue
    "ALINTA": "FFFFA500",       # Orange
    "NBE": "FF90EE90",          # Light Green (Next Business Energy)
    "ORIGIN": "FFFF0000",       # Red
    "1ST ENERGY": "FF00008B",   # Dark Blue
    "EA": "FF808080",           # Grey (Energy Australia)
}


def retailer_color_hex(retailer: str | None) -> str | None:
    if not retailer:
        return None
    return RETAILER_COLORS.get(retailer.strip().upper())


# ---------------------------------------------------------------------------
# Retailer selection rules (business logic, on top of the raw cost engine)
# ---------------------------------------------------------------------------
# Maps free-text retailer names as they show up on a bill (Gemini extraction
# of bill_data["current_energy_retailer"]) to the internal retailer keys used
# above. Anything that doesn't match one of these patterns (e.g. "AGL",
# "Powershop", "Simply Energy"...) returns None - it's simply not one of the
# retailers this sheet can propose, so no exclusion rule applies to it.
_RETAILER_NAME_PATTERNS = [
    ("ORIGIN", ("origin",)),
    ("MOMENTUM", ("momentum",)),
    ("1ST ENERGY", ("1st energy", "first energy")),
    ("NBE", ("next business", "nbe")),
    ("ALINTA", ("alinta",)),
    ("EA", ("energyaustralia", "energy australia")),
    ("GLOBIRD", ("globird", "glo bird")),
]


def normalize_retailer_name(raw: str | None) -> str | None:
    text = _norm(raw)
    if not text:
        return None
    for key, patterns in _RETAILER_NAME_PATTERNS:
        if any(p in text for p in patterns):
            return key
    return None


# Rule 1: never re-propose the customer's CURRENT retailer, with one
# exception - NBE and Origin are allowed to be re-proposed, but ONLY as a
# last resort when nothing else on the sheet actually saves the customer
# money. Every other retailer (Alinta, 1st Energy, EA, GloBird, or any
# retailer added to the sheet later) is excluded outright, saving or not.
REUSE_CURRENT_IF_NO_SAVING = {"NBE", "ORIGIN"}

# Rule 2: sites using this much or more per year are steered toward NBE,
# even if NBE isn't the single cheapest option on paper (commission /
# relationship reasons, not pure cost). Annual usage is estimated from the
# bill's usage-type charge quantities, scaled to a full year - see
# `_estimate_annual_kwh`.
HIGH_USAGE_KWH_THRESHOLD = 20000

# Rule 3: for tariffs classified as Residential, try these retailers first
# and only look wider if none of them beats the customer's current bill.
# NOTE: GloBird has NO columns in Master_Retailers_Comparison.xlsx yet (the
# sheet only has ORIGIN/MOMENTUM/1ST ENERGY/NBE/ALINTA/EA) - it's listed
# here for when that data exists, but until then it can never actually be
# selected. Flagged in meta["globird_unavailable"] whenever this matters.
RESIDENTIAL_PREFERRED_RETAILERS = ["1ST ENERGY", "ALINTA", "GLOBIRD"]

# Which Tariff Codes count as "Residential" for rule 3. Raj: this needs your
# actual list of Residential NTCs - populate it here (strings, matched
# case/space-insensitively against the live NTC), e.g.:
RESIDENTIAL_TARIFF_CODES = {"10", "26", "15", "16", "20", "30", "EA010", "EA111", "EA251", "EA210", "EA011", "EA116", "EA025_26", "EA025", "N70", "N72", "NS70", "N73", "N73N50", "N71_26", "N71", "N705", "N706", "BLNBSS1", "BLNN2AU", "BLNT3AU", "BLNRSS2", "BLND1AR", "BLNT3AL", "6900", "8700", "8900", "8920", "8950", "8970", "8400", "8420", "8450", "8470", "3900", "3970", "3950", "3750", "OPCL", "CL", "QRSR", "MRSR", "RSR", "RTOU", "NASS11", "NAST11", "NAST11S", "NAST11P", "NAST13", "NAST14", "NAST15", "SUN23", "NEE11", "NEE13", "NEE14", "NEE15", "NEE13CL", "C1R", "C1RB", "CR", "CRSTOU", "CRTOU", "A100", "F10I", "A10D", "A130", "A120", "F120", "D1", "DD", "PRSTOU", "PRTOU", "LVS1R", "RESKW1R", "URSTOU", "URTOU", "FURTOU"}
# Left empty for now, meaning rule 3 never fires until this is filled in.
RESIDENTIAL_TARIFF_CODES: set[str] = set()


def is_residential_tariff(ntc: str) -> bool:
    return _norm(ntc) in {_norm(c) for c in RESIDENTIAL_TARIFF_CODES}


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip().lower()


@dataclass
class TariffRow:
    row_num: int
    type_: str
    tariff_code: str
    cl_components: str
    state: str
    distribution: str
    retailers: dict = field(default_factory=dict)  # retailer -> {field: value|None}


#  In-process cache, keyed by (path, mtime), so the comparison workbook is
#  only ever parsed once per file version instead of on every single quote
#  generation. build_proposed_charges() -> load_comparison() used to be
#  called fresh (a full openpyxl parse of a wide 6-retailer x 14-col sheet)
#  for every single-site quote - on Streamlit's resource-limited hosting
#  tier this was the single biggest contributor to the container's memory
#  climbing over a session until it got OOM-killed (surfaces as the
#  "Segmentation fault" you saw, since the killed process can die mid
#  native-library call rather than with a clean Python traceback).
_comparison_cache: dict[str, tuple[float, list["TariffRow"]]] = {}


def load_comparison(path: str) -> list[TariffRow]:
    """Loads (and caches) every tariff row from the Comparison sheet.

    Cache is invalidated automatically if the file's mtime changes (e.g. you
    replace Master_Retailers_Comparison.xlsx with an updated version), so
    there's no need to restart the app after updating rates - just re-upload
    the file with a newer mtime.

    Uses openpyxl's `read_only=True` streaming mode rather than the default
    fully-materialized workbook: read_only mode never builds an in-memory
    object graph for the whole sheet, it streams row-by-row instead, which
    uses a fraction of the memory for a read-only extraction like this one.
    """
    mtime = os.path.getmtime(path)
    cached = _comparison_cache.get(path)
    if cached is not None and cached[0] == mtime:
        return cached[1]

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[COMPARISON_SHEET]
        rows = []
        # values_only=True skips building Cell objects entirely and just
        # hands back plain Python values - the cheapest possible read.
        for r, row_values in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
            tariff_code = row_values[1] if len(row_values) > 1 else None
            if tariff_code in (None, ""):
                continue
            retailers = {}
            for name, start_col in RETAILER_START_COL.items():
                vals = {}
                for i, fname in enumerate(FIELD_ORDER):
                    col_idx = start_col + i - 1  # 1-indexed sheet col -> 0-indexed tuple
                    raw = row_values[col_idx] if col_idx < len(row_values) else None
                    if fname in CENTS_TO_DOLLARS_FIELDS and isinstance(raw, (int, float)):
                        raw = raw / 100.0
                    vals[fname] = raw
                retailers[name] = vals
            rows.append(
                TariffRow(
                    row_num=r,
                    type_=row_values[0] if len(row_values) > 0 else None,
                    tariff_code=str(tariff_code).strip(),
                    cl_components=row_values[2] if len(row_values) > 2 else None,
                    state=row_values[3] if len(row_values) > 3 else None,
                    distribution=row_values[4] if len(row_values) > 4 else None,
                    retailers=retailers,
                )
            )
    finally:
        # read_only workbooks keep an open zip file handle until closed -
        # always release it, even if parsing raised.
        wb.close()

    _comparison_cache[path] = (mtime, rows)
    return rows


# ---------------------------------------------------------------------------
# Charge line -> rate field classification
# ---------------------------------------------------------------------------
def classify_charges(charges: list[dict]) -> list[str | None]:
    """
    Maps each bill charge line to one of CHARGE_FIELDS by keyword, in order.
    Handles the "second occurrence" ambiguity for peak/peak2 and
    demand1/demand2 by counting - the first "peak"-ish line found is `peak`,
    the next is `peak2`, etc. Returns a list the same length as `charges`,
    with None for any line that couldn't be confidently classified.
    """
    seen_peak = 0
    seen_demand = 0
    out: list[str | None] = []
    for c in charges:
        d = _norm(c.get("description"))
        u = _norm(c.get("unit"))
        text = f"{d} {u}"

        if "solar" in text or "feed" in text or "feed-in" in text or "fit" in d.split():
            out.append("solar")
        elif "daily" in text or ("supply" in text and "charge" in text):
            out.append("daily_supply")
        elif "shoulder" in text:
            out.append("shoulder")
        elif "off peak" in text or "offpeak" in text or "off-peak" in text:
            out.append("offpeak")
        elif "controlled load" in text or re.search(r"\bcl\s*1\b", text):
            out.append("cl1")
        elif re.search(r"\bcl\s*2\b", text):
            out.append("cl2")
        elif re.search(r"\bcl\s*3\b", text):
            out.append("cl3")
        elif "capacity" in text:
            out.append("capacity")
        elif "demand" in text:
            seen_demand += 1
            out.append("demand1" if seen_demand == 1 else "demand2")
        elif "peak" in text:  # catches plain "Peak Usage" after off-peak/shoulder are ruled out above
            seen_peak += 1
            out.append("peak" if seen_peak == 1 else "peak2")
        else:
            out.append(None)
    return out


# ---------------------------------------------------------------------------
# Matching + costing
# ---------------------------------------------------------------------------
def _find_candidates(rows: list[TariffRow], ntc: str, state: str | None, distribution: str | None):
    ntc_n = _norm(ntc)
    state_n = _norm(state)
    dist_n = _norm(distribution)

    by_code = [r for r in rows if _norm(r.tariff_code) == ntc_n]
    if not by_code:
        return [], "no_tariff_code_match"

    if state_n:
        by_code_state = [r for r in by_code if _norm(r.state) == state_n]
    else:
        by_code_state = by_code

    if state_n and dist_n:
        by_all = [r for r in by_code_state if _norm(r.distribution) == dist_n]
        if by_all:
            return by_all, "tariff_state_distribution"

    if by_code_state:
        return by_code_state, "tariff_state"

    return by_code, "tariff_code_only"


def _missing(value) -> bool:
    return value is None or value == ""


def _cost_for_retailer(charges: list[dict], field_map: list[str | None], rates: dict, billing_days: float):
    """
    Returns (total_cost, per_line_rates, unavailable) for one retailer's
    rate card against this bill's actual quantities.
    `unavailable` is True if any matched line has no published rate for that
    retailer on this tariff row (can't be fully priced -> disqualified).
    """
    discount_raw = rates.get("discount")
    discount_frac = (discount_raw / 100.0) if isinstance(discount_raw, (int, float)) else 0.0

    total = 0.0
    per_line = []
    unavailable = False
    for charge, fname in zip(charges, field_map):
        if fname is None:
            per_line.append(None)
            continue
        rate = rates.get(fname)
        if _missing(rate):
            unavailable = True
            per_line.append(None)
            continue
        qty = charge.get("quantity") or 0
        signed_rate = -abs(rate) if fname in CREDIT_FIELDS else rate
        line_cost = qty * signed_rate * (1 - discount_frac)
        total += line_cost
        per_line.append({"rate": signed_rate, "discount_pct": discount_frac})
    return total, per_line, unavailable


def _estimate_annual_kwh(charges: list[dict], field_map: list[str | None], billing_days: float) -> float | None:
    """
    Sums the bill's usage-type (kWh) charge quantities - Peak/Peak2/Shoulder/
    Off Peak/CL1-3, explicitly excluding Daily Supply, Capacity, Demand
    (kW not kWh) and Solar (generation credit, not consumption) - and scales
    the billing-period total up to a full year. Returns None if none of the
    bill's lines classified as a usage field (can't estimate).
    """
    usage_fields = {"peak", "peak2", "shoulder", "offpeak", "cl1", "cl2", "cl3"}
    total_kwh = 0.0
    found = False
    for charge, fname in zip(charges, field_map):
        if fname in usage_fields:
            total_kwh += charge.get("quantity") or 0
            found = True
    if not found:
        return None
    days = billing_days or 30
    return total_kwh * (365.0 / days)


def _eligible_retailers(candidates: list[TariffRow], charges: list[dict], field_map: list[str | None], billing_days: float):
    """
    Returns {retailer: (TariffRow, total_cost, per_line_rates)} for every
    retailer that can be FULLY priced (every matched charge line has a
    published rate) on at least one of the candidate tariff rows. If more
    than one candidate row prices the same retailer, the cheapest is kept.
    """
    best: dict[str, tuple] = {}
    for row in candidates:
        for retailer, rates in row.retailers.items():
            if all(_missing(v) for v in rates.values()):
                continue
            total, per_line, unavailable = _cost_for_retailer(charges, field_map, rates, billing_days)
            if unavailable:
                continue
            if retailer not in best or total < best[retailer][1]:
                best[retailer] = (row, total, per_line)
    return best


def _current_bill_total(charges: list[dict]) -> float | None:
    try:
        return sum(
            (c.get("quantity") or 0)
            * (c.get("rate_before_discount") or 0)
            * (1 - (c.get("conditional_discount_pct") or 0))
            * (-1 if c.get("is_credit") else 1)
            for c in charges
        )
    except Exception:
        return None





# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------
def select_offer(
    rows: list[TariffRow],
    charges: list[dict],
    ntc: str,
    state: str | None,
    distribution: str | None,
    current_retailer_raw: str | None,
    billing_days: float = 30,
):
    """
    Applies the retailer selection rules on top of the raw cost engine:

      1. Never re-propose the customer's current retailer, EXCEPT NBE and
         Origin, which may be re-proposed only if nothing else saves money.
      2. Sites estimated at 20,000+ kWh/year are steered toward NBE, even if
         it isn't the strict cheapest option.
      3. Residential-tariff bills (RESIDENTIAL_TARIFF_CODES) are steered
         toward 1st Energy / Alinta / GloBird first; if none of those beats
         the current bill, the search widens to every eligible retailer.
      4. Otherwise (or if the above tiers find nothing that saves money):
         cheapest eligible retailer that actually beats the current bill,
         excluding the current retailer.
      5. If literally nothing beats the current bill: NBE/Origin fall back
         to being reused as-is; any other current retailer instead gets the
         best-effort cheapest alternative on the sheet, clearly flagged in
         meta as not an actual saving (see `build_proposed_charges`).

    Returns a dict shaped like the old find_best_offer(), plus "tier" (which
    rule fired) and "current_retailer" (normalized key or None).
    """
    field_map = classify_charges(charges)
    candidates, match_level = _find_candidates(rows, ntc, state, distribution)
    eligible = _eligible_retailers(candidates, charges, field_map, billing_days)
    current_total = _current_bill_total(charges)
    current_key = normalize_retailer_name(current_retailer_raw)
    annual_kwh = _estimate_annual_kwh(charges, field_map, billing_days)
    residential = is_residential_tariff(ntc)

    # Rule 1: current retailer is excluded from the primary search pool.
    primary_pool = [r for r in eligible if r != current_key]

    def cheapest(pool):
        items = [(r, eligible[r][1]) for r in pool if r in eligible]
        return min(items, key=lambda x: x[1])[0] if items else None

    def saves_money(retailer):
        return current_total is None or eligible[retailer][1] < current_total

    chosen, tier = None, None

    # Rule 2: high usage -> prefer NBE outright if it's a candidate at all.
    if annual_kwh is not None and annual_kwh >= HIGH_USAGE_KWH_THRESHOLD and "NBE" in primary_pool:
        chosen, tier = "NBE", f"high_usage_prefers_nbe (~{annual_kwh:,.0f} kWh/yr est.)"

    # Rule 3: residential tariff -> try the preferred group first.
    if chosen is None and residential:
        pref_pool = [r for r in primary_pool if r in RESIDENTIAL_PREFERRED_RETAILERS]
        candidate = cheapest(pref_pool)
        if candidate is not None and saves_money(candidate):
            chosen, tier = candidate, "residential_tariff_preferred"

    # Default: cheapest retailer that actually beats the current bill.
    if chosen is None:
        candidate = cheapest(primary_pool)
        if candidate is not None and saves_money(candidate):
            chosen, tier = candidate, "cheapest_saving"

    # Rule 1 exception + best-effort fallback: nothing above saved money.
    if chosen is None:
        if current_key in REUSE_CURRENT_IF_NO_SAVING and current_key in eligible:
            chosen, tier = current_key, "no_saving_reuse_current"
        else:
            candidate = cheapest(primary_pool)
            if candidate is not None:
                chosen, tier = candidate, "no_saving_best_effort"

    globird_unavailable = residential and "GLOBIRD" not in eligible and "GLOBIRD" not in RETAILER_START_COL

    if chosen is None:
        return {
            "retailer": None,
            "tariff_row": None,
            "total_cost": None,
            "per_line_rates": [None] * len(charges),
            "match_level": match_level,
            "field_map": field_map,
            "candidates_considered": len(candidates),
            "eligible_retailers": sorted(eligible.keys()),
            "tier": tier or "no_eligible_retailer",
            "current_retailer": current_key,
            "current_total": current_total,
            "residential": residential,
            "annual_kwh_estimate": annual_kwh,
            "globird_unavailable": globird_unavailable,
        }

    row, total, per_line = eligible[chosen]
    return {
        "retailer": chosen,
        "tariff_row": row,
        "total_cost": total,
        "per_line_rates": per_line,
        "match_level": match_level,
        "field_map": field_map,
        "candidates_considered": len(candidates),
        "eligible_retailers": sorted(eligible.keys()),
        "signup_credit": row.retailers[chosen].get("signup_credit"),
        "tier": tier,
        "current_retailer": current_key,
        "current_total": current_total,
        "residential": residential,
        "annual_kwh_estimate": annual_kwh,
        "globird_unavailable": globird_unavailable,
    }
def build_proposed_charges(bill_data: dict, ntc: str, comparison_path: str) -> tuple[list[dict], dict]:
    """
    High-level helper for app.py:

        proposed_charges, meta = build_proposed_charges(
            extracted["data"], live_ntc, "Master_Retailers_Comparison.xlsx"
        )
        fill_quote(extracted["data"], out_path,
                   tariff_override=live_ntc,
                   proposed_charges=proposed_charges,
                   proposed_retailer_color=retailer_color_hex(meta["retailer"]))

    `proposed_charges` is the SAME length/order as bill_data["charges"], each
    entry either None (leave that row's J/K blank - couldn't price it) or
    {"rate_before_discount": .., "conditional_discount_pct": .., "is_credit": ..}
    ready to be written by single_excel_filler in place of the bill's own
    rate_before_discount/conditional_discount_pct for that row.

    Retailer choice applies the rules in select_offer() - see its docstring
    - not just "cheapest wins".
    """
    charges = bill_data.get("charges") or []
    rows = load_comparison(comparison_path)
    result = select_offer(
        rows,
        charges,
        ntc=ntc,
        state=bill_data.get("state"),
        distribution=bill_data.get("distribution_region"),
        current_retailer_raw=bill_data.get("current_energy_retailer"),
        billing_days=bill_data.get("billing_period_days") or 30,
    )

    proposed_charges = []
    for charge, line in zip(charges, result["per_line_rates"]):
        if line is None:
            proposed_charges.append(None)
        else:
            proposed_charges.append(
                {
                    "rate_before_discount": abs(line["rate"]),
                    "conditional_discount_pct": line["discount_pct"],
                    "is_credit": line["rate"] < 0,
                }
            )

    unmatched = [
        c.get("description")
        for c, f in zip(charges, result["field_map"])
        if f is None
    ]

    current_total = result.get("current_total")
    proposed_total = result.get("total_cost")
    saving = (
        current_total - proposed_total
        if current_total is not None and proposed_total is not None
        else None
    )

    meta = {
        "retailer": result["retailer"],
        "tariff_code": result["tariff_row"].tariff_code if result["tariff_row"] else None,
        "match_level": result["match_level"],
        "candidates_considered": result["candidates_considered"],
        "eligible_retailers": result["eligible_retailers"],
        "proposed_total": proposed_total,
        "current_total": current_total,
        "estimated_saving": saving,
        # True only when the chosen retailer was a best-effort pick that
        # does NOT actually beat the current bill (tier
        # "no_saving_reuse_current" / "no_saving_best_effort") - the UI
        # should say so rather than implying a saving that isn't there.
        "no_actual_saving": bool(saving is not None and saving <= 0),
        "selection_tier": result.get("tier"),
        "current_retailer_normalized": result.get("current_retailer"),
        "is_residential_tariff": result.get("residential"),
        "annual_kwh_estimate": result.get("annual_kwh_estimate"),
        "globird_unavailable": result.get("globird_unavailable"),
        "signup_credit": result.get("signup_credit"),
        "unmatched_descriptions": unmatched,
    }
    return proposed_charges, meta
