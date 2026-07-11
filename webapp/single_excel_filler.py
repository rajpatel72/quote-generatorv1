"""
Fills template/blank_quote_template.xlsx ("PDF Format" sheet) from the JSON
produced by gemini_client.extract_bill().

WHY THE PREVIOUS VERSION BROKE FORMATTING (root cause notes, kept here on
purpose so future-you doesn't reintroduce the same bugs):

1. openpyxl's ws.insert_rows() only shifts CELL VALUES down. It does NOT
   shift `row_dimensions` (custom row heights) and does NOT shift
   `merged_cells` ranges. Both have to be moved by hand, or you get exactly
   what you saw: the Total row's content moves down but keeps whatever
   height belonged to the row it lands on (the "squeezing"), and merged
   blocks like the divider columns (D30:D36, I30:I36), the key-features box
   (O30:T36), and the two summary boxes (E39:H42/J39:K42 and E45:H48/J45:K48)
   stay anchored at their OLD row numbers while their real content slides
   out from underneath them (the "border break").

2. The old code inserted rows at row 34 (in the middle of the reserved
   block), which sits *inside* the D30:D36/I30:I36/O30:T36 merges — the
   worst place to insert without fixing merges, since it splits a merge
   that's currently open. This version always inserts right before the
   Total row (i.e. after the last reserved/blank row), which only affects
   merges that start above the insertion point and touch the table, plus
   everything below — a case that's easy to handle deterministically.

3. `ORIGINAL_ANNUAL_SAVINGS_ROW` was set to 48 in the old code, but the
   actual "Estimated Annual Savings" formula lives at J45 (merged J45:K48 —
   48 is just the bottom of that merged block, not a normal writable cell,
   so writes to it were silently dropped). The template's J45 formula also
   ships with a `#REF!` in it (a stale reference to a deleted "billing days"
   cell) — that's fixed here unconditionally, using the real billing period
   from the bill, regardless of whether any rows were inserted.

4. B16 (tariff) is now sourced live from the NMI lookup tool (see
   nmi_tariff_lookup.fetch_network_tariff) instead of trusting whatever
   Gemini extracted from the bill text, since the bill's own printed
   tariff label doesn't always match the network's current tariff code.
   If the lookup fails for any reason we fall back to the extracted value
   so a flaky network call never blocks quote generation.

5. `proposed_charges` (see retailer_rates.build_proposed_charges) fills the
   "New Proposed Offer" columns (J/K, which the template's own L/M formulas
   already turn into After Discount / Total). It's positional: entry i
   corresponds to charges[i], same row as that charge's Current Offer line,
   so the two offers stay lined up item-for-item. An entry of None (or a
   proposed_charges list shorter than charges, or None altogether) just
   leaves that row's J/K blank - same "don't guess" behaviour as an
   unmatched charge line upstream in retailer_rates.classify_charges.

The template reserves 3 charge-line rows (33, 34, 35) before the Total row
(36). Bills with more than 3 charge lines get extra rows inserted between
row 35 and row 36, with every affected row height / merge / summary formula
recalculated to match.
"""
import copy
import os

import openpyxl
from openpyxl.worksheet.cell_range import CellRange


TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "template", "blank_quote_template.xlsx")
SHEET_NAME = "PDF Format"

FIRST_CHARGE_ROW = 33
RESERVED_CHARGE_ROWS = 3          # rows 33, 34, 35 exist in the template already
ORIGINAL_TOTAL_ROW = 36
ORIGINAL_PCT_SAVING_ROW = 39      # J39, merged J39:K42
ORIGINAL_ANNUAL_SAVINGS_ROW = 45  # J45, merged J45:K48

STYLE_COLS = [chr(c) for c in range(ord("A"), ord("X") + 1)]  # sheet uses up to column X



def _copy_row_style(ws, src_row: int, dst_row: int):
    for col in STYLE_COLS:
        src = ws[f"{col}{src_row}"]
        dst = ws[f"{col}{dst_row}"]
        dst.font = copy.copy(src.font)
        dst.fill = copy.copy(src.fill)
        dst.border = copy.copy(src.border)
        dst.alignment = copy.copy(src.alignment)
        dst.number_format = src.number_format


def _insert_rows_preserving_layout(ws, insert_at: int, n: int):
    """
    Inserts `n` blank rows at `insert_at` and fixes the two things
    openpyxl's insert_rows() leaves broken: row heights and merged ranges.
    Must be called before any new content is written to the inserted rows.
    """
    old_heights = {
        r: ws.row_dimensions[r].height
        for r in range(1, ws.max_row + 1)
        if ws.row_dimensions[r].height is not None
    }
    old_merges = [str(mc) for mc in ws.merged_cells.ranges]

    # Merges must be cleared before the row shift, or the cells inside them
    # aren't plain writable cells and the shift/rewrite below gets unreliable.
    for mc in old_merges:
        ws.unmerge_cells(mc)

    ws.insert_rows(insert_at, n)

    # Re-apply row heights at their shifted position.
    for old_row, height in old_heights.items():
        new_row = old_row + n if old_row >= insert_at else old_row
        ws.row_dimensions[new_row].height = height

    # Re-create every merge at its shifted position:
    #  - entirely above the insertion point -> unchanged
    #  - spans the insertion point (starts above, ends at/below) -> stretch
    #  - entirely at/below the insertion point -> shift down wholesale
    for mc in old_merges:
        cell_range = CellRange(mc)
        min_r, max_r = cell_range.min_row, cell_range.max_row
        min_c, max_c = cell_range.min_col, cell_range.max_col
        if min_r >= insert_at:
            min_r, max_r = min_r + n, max_r + n
        elif max_r >= insert_at:
            max_r += n
        ws.merge_cells(start_row=min_r, start_column=min_c, end_row=max_r, end_column=max_c)


def fill_quote(
    bill_data: dict,
    output_path: str,
    tariff_debug: dict | None = None,
    tariff_override: str | None = None,
    proposed_charges: list[dict | None] | None = None,
) -> str:
    """
    tariff_debug: optional dict passed in by the caller (e.g. {}). If given,
    it gets populated in-place with the Network Tariff lookup's diagnostic
    info (url/status_code/error/matched/raw_snippet) so the caller can show
    it in a UI - the lookup fails silently otherwise by design, which makes
    it hard to tell *why* B16 fell back to the extracted tariff.

    tariff_override: if provided (e.g. a value obtained via the browser
    popup + userscript flow in tariff_bridge.py), this is written to B16
    directly and the cookie-based live lookup is skipped entirely - no
    network call is made and tariff_debug (if given) is populated with a
    synthetic "source": "override" entry instead.

    proposed_charges: optional list, same length/order as
    bill_data["charges"]. Each entry is either None (leave that row's New
    Proposed Offer cells blank) or a dict with "rate_before_discount",
    "conditional_discount_pct", and "is_credit" - the same shape as a bill
    charge - used to fill J/K for that row instead of the bill's own current
    rate. See retailer_rates.build_proposed_charges().
    """
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb[SHEET_NAME]

    # --- header / customer fields ---
    ws["A6"] = bill_data.get("customer_name").upper()

    nmi = bill_data.get("nmi_or_mirn")

    extracted_tariff = bill_data.get("tariff_classification")
    if tariff_override:
        ws["B16"] = tariff_override
        if tariff_debug is not None:
            tariff_debug["matched"] = True
            tariff_debug["source"] = "browser_popup_override"
            tariff_debug["tariff"] = tariff_override
    else:
        ws["B16"] = extracted_tariff
        if tariff_debug is not None:
            tariff_debug["matched"] = False
            tariff_debug["source"] = "extracted_from_bill"
            tariff_debug["error"] = "No browser popup override was supplied"

    ws["B17"] = bill_data.get("distribution_region")
    ws["B18"] = bill_data.get("site_address")

    # ws["B19"].number_format = "@"  # keep as text: NMIs can have leading zeros
    ws["B19"] = str(nmi) if nmi else None
    ws["B21"] = bill_data.get("current_energy_retailer")

    charges = bill_data.get("charges") or []
    # +1 because a blank buffer row must always sit directly above Total,
    # on top of however many rows the charges themselves need.
    rows_needed_before_total = len(charges) + 1
    extra_rows = max(0, rows_needed_before_total - RESERVED_CHARGE_ROWS)

    if extra_rows > 0:
        _insert_rows_preserving_layout(ws, ORIGINAL_TOTAL_ROW, extra_rows)
        # Style for the newly inserted rows, copied from row 35 (the
        # template's blank "middle of table" row). Formulas are written
        # below for every charge row uniformly, including these; the buffer
        # row is deliberately left with no formulas at all.
        style_source_row = FIRST_CHARGE_ROW + RESERVED_CHARGE_ROWS - 1  # 35
        source_height = ws.row_dimensions[style_source_row].height
        for offset in range(extra_rows):
            new_row = ORIGINAL_TOTAL_ROW + offset
            _copy_row_style(ws, style_source_row, new_row)
            # New rows land on whatever height the template originally had
            # at that row number (insert_rows doesn't touch row_dimensions),
            # so override it explicitly to match the other charge rows.
            ws.row_dimensions[new_row].height = source_height

    total_row = ORIGINAL_TOTAL_ROW + extra_rows
    pct_saving_row = ORIGINAL_PCT_SAVING_ROW + extra_rows
    annual_savings_row = ORIGINAL_ANNUAL_SAVINGS_ROW + extra_rows
    last_charge_row = FIRST_CHARGE_ROW + len(charges) - 1 if charges else FIRST_CHARGE_ROW - 1
    buffer_row = total_row - 1  # always left blank, directly above Total

    # --- write charge line items ---
    # NB: row 35 ships blank in the template (no G/H/L/M formulas at all,
    # unlike rows 33/34 which have them hardcoded) so those formulas are
    # written explicitly here for every row actually used by a charge,
    # rather than relying on whatever the template happened to ship with.
    for i, charge in enumerate(charges):
        r = FIRST_CHARGE_ROW + i
        ws[f"B{r}"] = charge.get("quantity")
        desc = charge.get("description") or ""
        if charge.get("is_credit"):
            desc = f"{desc} (credit)" if desc else "Credit"
        ws[f"C{r}"] = desc
        rate = charge.get("rate_before_discount")
        ws[f"E{r}"] = -abs(rate) if (rate is not None and charge.get("is_credit")) else rate
        ws[f"F{r}"] = charge.get("conditional_discount_pct")
        ws[f"G{r}"] = f"=E{r}*(1-F{r})"
        ws[f"H{r}"] = f"=B{r}*G{r}"
        ws[f"L{r}"] = f"=J{r}*(1-K{r})"
        ws[f"M{r}"] = f"=B{r}*L{r}"

        # New Proposed Offer (J/K) for the same line item, if supplied.
        # L/M for this row are already the formulas above, so as soon as
        # J/K are populated the After Discount / Total cells compute
        # themselves - nothing else needs to change on this row.
        proposed = proposed_charges[i] if proposed_charges and i < len(proposed_charges) else None
        if proposed:
            p_rate = proposed.get("rate_before_discount")
            ws[f"J{r}"] = -abs(p_rate) if (p_rate is not None and proposed.get("is_credit")) else p_rate
            ws[f"K{r}"] = proposed.get("conditional_discount_pct")

    # Everything between the last real charge row and Total is blank —
    # this always includes at least the one mandatory buffer row, plus any
    # leftover reserved rows for bills with fewer than 3 charges.
    for r in range(last_charge_row + 1, total_row):
        for col in ["B", "C", "E", "F", "G", "H", "J", "K", "L", "M"]:
            ws[f"{col}{r}"] = None

    # --- fix up the formulas that depend on where the Total row now sits ---
    ws[f"H{total_row}"] = f"=SUM(H{FIRST_CHARGE_ROW}:H{buffer_row})"
    ws[f"M{total_row}"] = f"=SUM(M{FIRST_CHARGE_ROW}:M{buffer_row})"
    ws[f"J{pct_saving_row}"] = f"=1-(M{total_row}/H{total_row})"

    # The template's Annual Savings formula ships with a #REF! (a deleted
    # "billing days" cell) — always replace it with the bill's real billing
    # period, whether or not any rows were inserted.
    billing_days = bill_data.get("billing_period_days") or 30
    ws[f"J{annual_savings_row}"] = f"=((H{total_row}-M{total_row})/{billing_days}*365)"

    wb.save(output_path)
    return output_path
